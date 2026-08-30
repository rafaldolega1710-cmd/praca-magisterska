"""Testy pipeline'u danych (`src/data_loader.py`).

Żadna z tych testów nie wykonuje prawdziwych zapytań sieciowych -- HTTP jest
zamockowane, a parsowanie plików binarnych (xls) testowane na małym fixture
wygenerowanym w locie. To celowe: testy muszą przechodzić identycznie u
każdego, bez dostępu do internetu i bez zależności od tego, czy zewnętrzne
API akurat odpowiada.
"""

from unittest.mock import MagicMock, patch

import openpyxl
import pandas as pd
import pytest

from src.data_loader import (
    _fetch_nbp_daily_for_year,
    _nbp_reference_rate_at,
    _parse_damodaran_xls,
    _parse_nbp_archive_year,
    _parse_price_series_to_monthly_returns,
    annualize_to_monthly,
    build_edo_reference_rate_monthly,
    fetch_gus_series,
    fetch_nbp_reference_rate,
    fetch_nbp_usdpln_monthly,
    load_acwi_history,
    load_edo_margins,
)


class TestAnnualizeToMonthly:
    def test_zero_annual_return_is_zero_monthly(self):
        assert annualize_to_monthly(0.0) == pytest.approx(0.0)

    def test_known_compounding_value(self):
        # (1.01)**12 - 1 ~= 0.126825 -- zaokrąglona roczna stopa odtwarzająca
        # dokładnie 1% miesięcznie po złożeniu
        annual = 1.01**12 - 1
        assert annualize_to_monthly(annual) == pytest.approx(0.01)

    def test_negative_annual_return(self):
        # spadek o 50% w skali roku -> ujemna stopa miesięczna
        result = annualize_to_monthly(-0.5)
        assert result < 0.0
        assert (1 + result) ** 12 == pytest.approx(0.5)


class TestParseDamodaranXls:
    def _make_fixture(self, tmp_path):
        """Buduje minimalny plik xls odwzorowujący zweryfikowany układ
        arkusza 'Returns by year': 19 wierszy nagłówkowych/notatek, nagłówek
        kolumn w wierszu 20 (1-indeksowanym), potem dane roczne."""
        path = tmp_path / "histretSP_fixture.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Returns by year"
        for _ in range(19):  # wiersze 1..19: notatki/metadane, jak w realnym pliku
            ws.append([None])
        ws.append(
            [
                "Year",
                "S&P 500 (includes dividends)",
                "US Small cap (bottom decile)",
                "3-month T.Bill",
                "US T. Bond (10-year)",
            ]
        )
        ws.append([2020, 0.1840, 0.20, 0.0054, 0.1133])
        ws.append([2021, 0.2871, 0.25, 0.0005, -0.0404])
        ws.append([2022, -0.1811, -0.20, 0.0201, -0.1259])
        wb.save(path)
        return path

    def test_parses_expected_columns_and_index(self, tmp_path):
        path = self._make_fixture(tmp_path)
        df = _parse_damodaran_xls(path)
        assert list(df.index) == [2020, 2021, 2022]
        assert df.loc[2021, "sp500_annual_return"] == pytest.approx(0.2871)
        assert df.loc[2022, "ust10y_annual_return"] == pytest.approx(-0.1259)


class TestFetchNbpUsdplnMonthly:
    def test_rejects_start_year_before_archive_range(self):
        with pytest.raises(ValueError, match="1995"):
            fetch_nbp_usdpln_monthly(start_year=1990)

    def test_resamples_daily_rates_to_monthly_last(self, tmp_path):
        cache_path = tmp_path / "nbp_usdpln.csv"

        mock_response = MagicMock()
        mock_response.raise_for_status.return_value = None
        mock_response.json.return_value = {
            "rates": [
                {"effectiveDate": "2002-01-10", "mid": 4.00},
                {"effectiveDate": "2002-01-31", "mid": 4.10},
                {"effectiveDate": "2002-02-15", "mid": 4.20},
                {"effectiveDate": "2002-02-28", "mid": 4.25},
            ]
        }

        with patch("src.data_loader.requests.get", return_value=mock_response):
            result = fetch_nbp_usdpln_monthly(
                start_year=2002, end_year=2002, cache_path=cache_path
            )

        assert result.loc[pd.Period("2002-01", freq="M"), "usd_pln"] == pytest.approx(4.10)
        assert result.loc[pd.Period("2002-02", freq="M"), "usd_pln"] == pytest.approx(4.25)

    def test_current_year_range_capped_at_today_not_dec_31(self, tmp_path):
        # regresja: zapytanie o pelny rok biezacy do 31 grudnia zwraca z NBP
        # API blad 400 "Invalid date range" dla dni, ktore jeszcze nie
        # nastapily -- zakres musi byc ucinany do dzisiejszej daty.
        cache_path = tmp_path / "nbp_usdpln.csv"
        current_year = pd.Timestamp.today().year

        mock_response = MagicMock()
        mock_response.raise_for_status.return_value = None
        mock_response.json.return_value = {"rates": []}

        with patch("src.data_loader.requests.get", return_value=mock_response) as mock_get:
            fetch_nbp_usdpln_monthly(
                start_year=current_year, end_year=current_year, cache_path=cache_path
            )

        called_url = mock_get.call_args[0][0]
        assert f"{current_year}-12-31" not in called_url

    def test_uses_cache_without_hitting_network(self, tmp_path):
        cache_path = tmp_path / "nbp_usdpln.csv"
        cache_path.write_text("date,usd_pln\n2002-01-31,4.10\n2002-02-28,4.25\n")

        with patch("src.data_loader.requests.get") as mock_get:
            result = fetch_nbp_usdpln_monthly(
                start_year=2002, end_year=2002, cache_path=cache_path
            )
            mock_get.assert_not_called()

        assert result.loc[pd.Period("2002-01", freq="M"), "usd_pln"] == pytest.approx(4.10)


class TestParseNbpArchiveYear:
    def _make_fixture(self, tmp_path, header_row, usd_header, rows):
        path = tmp_path / "archiwum_tab_a_test.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        for _ in range(header_row):
            ws.append([None])
        ws.append(["Nr / No.", "Data / Date", usd_header, "1 DEM"])
        for row in rows:
            ws.append(row)
        wb.save(path)
        return path

    def test_parses_1_usd_header_format(self, tmp_path):
        # uklad z lat 1996+
        path = self._make_fixture(
            tmp_path,
            header_row=1,
            usd_header="1 USD",
            rows=[
                [1, pd.Timestamp("1998-01-02"), 3.525, 1.956],
                [2, pd.Timestamp("1998-01-05"), 3.5275, 1.946],
            ],
        )
        df = _parse_nbp_archive_year(path)
        assert list(df["usd_pln"]) == pytest.approx([3.525, 3.5275])

    def test_parses_100_usd_header_format_and_divides_by_100(self, tmp_path):
        # uklad specyficzny dla 1995 r. -- kurs podany za 100 USD
        path = self._make_fixture(
            tmp_path,
            header_row=1,
            usd_header="100 USD",
            rows=[
                [1, pd.Timestamp("1995-01-02"), 243.01, 156.9],
                [2, pd.Timestamp("1995-01-03"), 243.64, 156.66],
            ],
        )
        df = _parse_nbp_archive_year(path)
        assert list(df["usd_pln"]) == pytest.approx([2.4301, 2.4364])

    def test_missing_usd_column_raises(self, tmp_path):
        path = tmp_path / "bad.xlsx"
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["Nr", "Data / Date", "1 EUR"])
        ws.append([1, pd.Timestamp("2000-01-01"), 4.0])
        wb.save(path)
        with pytest.raises(ValueError):
            _parse_nbp_archive_year(path)


class TestFetchNbpDailyForYear:
    def test_pre_2002_year_uses_archive_not_live_api(self, tmp_path, monkeypatch):
        monkeypatch.setattr("src.data_loader.NBP_ARCHIVE_DIR", tmp_path)
        fake_df = pd.DataFrame({"date": [pd.Timestamp("1999-01-04")], "usd_pln": [3.5]})

        with patch("src.data_loader._download_nbp_archive_year") as mock_download, \
             patch("src.data_loader._parse_nbp_archive_year", return_value=fake_df) as mock_parse, \
             patch("src.data_loader.requests.get") as mock_get:
            result = _fetch_nbp_daily_for_year(1999)
            mock_download.assert_called_once()
            mock_parse.assert_called_once()
            mock_get.assert_not_called()

        assert result.equals(fake_df)

    def test_2002_and_later_uses_live_api_not_archive(self):
        mock_response = MagicMock()
        mock_response.raise_for_status.return_value = None
        mock_response.json.return_value = {"rates": [{"effectiveDate": "2002-01-02", "mid": 4.0}]}

        with patch("src.data_loader.requests.get", return_value=mock_response) as mock_get, \
             patch("src.data_loader._download_nbp_archive_year") as mock_download:
            _fetch_nbp_daily_for_year(2002)
            mock_get.assert_called_once()
            mock_download.assert_not_called()


class TestFetchGusSeries:
    def test_parses_gus_bdl_response(self, tmp_path):
        cache_path = tmp_path / "gus_test.csv"
        mock_response = MagicMock()
        mock_response.raise_for_status.return_value = None
        mock_response.json.return_value = {
            "results": [
                {
                    "id": "000000000000",
                    "name": "POLSKA",
                    "values": [
                        {"year": "2022", "val": 114.4, "attrId": 1},
                        {"year": "2023", "val": 111.4, "attrId": 1},
                    ],
                }
            ]
        }

        with patch("src.data_loader.requests.get", return_value=mock_response):
            df = fetch_gus_series(217230, cache_path)

        assert df.loc[2022, "value"] == pytest.approx(114.4)
        assert df.loc[2023, "value"] == pytest.approx(111.4)
        assert cache_path.exists()  # wynik zapisany do cache

    def test_second_call_reads_cache_not_network(self, tmp_path):
        cache_path = tmp_path / "gus_test.csv"
        cache_path.write_text("year,value\n2022,114.4\n2023,111.4\n")

        with patch("src.data_loader.requests.get") as mock_get:
            df = fetch_gus_series(217230, cache_path)
            mock_get.assert_not_called()

        assert df.loc[2023, "value"] == pytest.approx(111.4)


class TestParsePriceSeriesToMonthlyReturns:
    def test_missing_file_raises_with_instructions(self, tmp_path):
        missing = tmp_path / "wig.csv"
        with pytest.raises(FileNotFoundError, match="README"):
            _parse_price_series_to_monthly_returns(missing, "wig_monthly_return")

    def test_parses_stooq_polish_format(self, tmp_path):
        path = tmp_path / "wig.csv"
        path.write_text(
            "Data;Otwarcie;Najwyzszy;Najnizszy;Zamkniecie;Wolumen\n"
            "20230131;1000;1050;990;1000;123456\n"
            "20230228;1000;1100;995;1100;123456\n"
            "20230331;1100;1150;1080;1210;123456\n",
            encoding="utf-8",
        )
        result = _parse_price_series_to_monthly_returns(path, "wig_monthly_return")

        assert result.loc[pd.Period("2023-02", freq="M"), "wig_monthly_return"] == pytest.approx(0.10)
        assert result.loc[pd.Period("2023-03", freq="M"), "wig_monthly_return"] == pytest.approx(0.10)
        # pierwszy miesiąc nie ma poprzedniej wartości do policzenia zwrotu
        assert pd.Period("2023-01", freq="M") not in result.index

    def test_parses_english_format(self, tmp_path):
        path = tmp_path / "tbsp.csv"
        path.write_text(
            "Date,Open,High,Low,Close,Volume\n"
            "2023-01-31,100,105,99,100,1000\n"
            "2023-02-28,100,110,99.5,110,1000\n",
            encoding="utf-8",
        )
        result = _parse_price_series_to_monthly_returns(path, "tbsp_monthly_return")
        assert result.loc[pd.Period("2023-02", freq="M"), "tbsp_monthly_return"] == pytest.approx(0.10)

    def test_unrecognized_columns_raise_value_error(self, tmp_path):
        path = tmp_path / "bad.csv"
        path.write_text("foo,bar\n1,2\n", encoding="utf-8")
        with pytest.raises(ValueError):
            _parse_price_series_to_monthly_returns(path, "x")


class TestLoadAcwiHistory:
    def test_missing_file_raises_with_readme_pointer(self, tmp_path):
        missing = tmp_path / "acwi_monthly.csv"
        with pytest.raises(FileNotFoundError, match="README"):
            load_acwi_history(missing)

    def test_computes_monthly_returns_from_index_level(self, tmp_path):
        path = tmp_path / "acwi_monthly.csv"
        path.write_text(
            "month,index_level\n2020-01,100.0\n2020-02,110.0\n2020-03,99.0\n",
            encoding="utf-8",
        )
        result = load_acwi_history(path)
        assert result.loc[pd.Period("2020-02", freq="M"), "acwi_monthly_return"] == pytest.approx(0.10)
        assert result.loc[pd.Period("2020-03", freq="M"), "acwi_monthly_return"] == pytest.approx(-0.10)
        # pierwszy miesiac nie ma poprzedniej wartosci do policzenia zwrotu
        assert pd.Period("2020-01", freq="M") not in result.index


class TestNbpReferenceRate:
    def test_missing_file_raises(self, tmp_path):
        with pytest.raises(FileNotFoundError):
            fetch_nbp_reference_rate(tmp_path / "missing.csv")

    def test_loads_and_sorts_by_date(self, tmp_path):
        path = tmp_path / "nbp_ref.csv"
        path.write_text(
            "effective_from,reference_rate_pct\n2020-06-01,1.00\n2019-01-01,1.50\n",
            encoding="utf-8",
        )
        df = fetch_nbp_reference_rate(path)
        assert list(df.index) == sorted(df.index)

    def test_rate_at_picks_last_value_on_or_before_month_end(self, tmp_path):
        path = tmp_path / "nbp_ref.csv"
        path.write_text(
            "effective_from,reference_rate_pct\n"
            "2019-01-01,1.50\n"
            "2020-06-15,1.00\n"
            "2021-01-01,2.00\n",
            encoding="utf-8",
        )
        df = fetch_nbp_reference_rate(path)
        # czerwiec 2020: zmiana wchodzi w zycie 15.06, wiec obowiazuje do konca miesiaca
        assert _nbp_reference_rate_at(pd.Period("2020-06", freq="M"), df) == pytest.approx(0.01)
        # maj 2020: jeszcze stara stawka z 2019
        assert _nbp_reference_rate_at(pd.Period("2020-05", freq="M"), df) == pytest.approx(0.015)

    def test_rate_before_first_entry_raises(self, tmp_path):
        path = tmp_path / "nbp_ref.csv"
        path.write_text("effective_from,reference_rate_pct\n2019-01-01,1.50\n", encoding="utf-8")
        df = fetch_nbp_reference_rate(path)
        with pytest.raises(ValueError):
            _nbp_reference_rate_at(pd.Period("2018-01", freq="M"), df)


class TestLoadEdoMargins:
    def test_missing_file_raises(self, tmp_path):
        with pytest.raises(FileNotFoundError):
            load_edo_margins(tmp_path / "missing.csv")

    def test_loads_indexed_by_issuance_month(self, tmp_path):
        path = tmp_path / "edo.csv"
        path.write_text(
            "issuance_month,first_year_rate_pct,margin_pct\n"
            "2013-09,,\n"
            "2017-01,,1.5\n",
            encoding="utf-8",
        )
        df = load_edo_margins(path)
        assert pd.isna(df.loc["2013-09", "margin_pct"])
        assert df.loc["2017-01", "margin_pct"] == pytest.approx(1.5)


class TestBuildEdoReferenceRateMonthly:
    def test_uses_cpi_plus_margin_when_margin_known(self, tmp_path):
        margins_path = tmp_path / "edo.csv"
        margins_path.write_text(
            "issuance_month,first_year_rate_pct,margin_pct\n2020-01,,1.5\n", encoding="utf-8"
        )
        nbp_path = tmp_path / "nbp.csv"
        nbp_path.write_text("effective_from,reference_rate_pct\n1998-01-01,10.0\n", encoding="utf-8")

        fake_cpi = pd.DataFrame({"cpi_prev_year_100": [103.4]}, index=pd.Index([2020], name="year"))
        with patch("src.data_loader.fetch_gus_cpi", return_value=fake_cpi):
            result = build_edo_reference_rate_monthly(margins_path, nbp_path)

        # inflacja 2020 = 3.4%, marza 1.5% -> stopa roczna 4.9%, przeliczona na miesieczna
        expected_annual = 0.034 + 0.015
        expected_monthly = annualize_to_monthly(expected_annual)
        assert result.loc[pd.Period("2020-01", freq="M"), "edo_reference_monthly_return"] == pytest.approx(
            expected_monthly
        )

    def test_falls_back_to_nbp_reference_plus_2pp_when_margin_unknown(self, tmp_path):
        margins_path = tmp_path / "edo.csv"
        # brak marzy dla 2015-06 (okres miedzy startem EDO a styczniem 2017)
        margins_path.write_text(
            "issuance_month,first_year_rate_pct,margin_pct\n2015-06,,\n", encoding="utf-8"
        )
        nbp_path = tmp_path / "nbp.csv"
        nbp_path.write_text(
            "effective_from,reference_rate_pct\n2015-01-01,1.50\n", encoding="utf-8"
        )

        fake_cpi = pd.DataFrame({"cpi_prev_year_100": [100.0]}, index=pd.Index([2015], name="year"))
        with patch("src.data_loader.fetch_gus_cpi", return_value=fake_cpi):
            result = build_edo_reference_rate_monthly(margins_path, nbp_path)

        # stopa ref. NBP 1.5% + fallback marzy 2% = 3.5% rocznie
        expected_monthly = annualize_to_monthly(0.015 + 0.02)
        assert result.loc[pd.Period("2015-06", freq="M"), "edo_reference_monthly_return"] == pytest.approx(
            expected_monthly
        )

    def test_falls_back_when_margin_known_but_cpi_not_yet_published(self, tmp_path):
        # regresja: rok biezacy (np. 2026) moze miec znana marze EDO, ale GUS
        # jeszcze nie opublikowal CPI za ten rok -- formula EDO jest wtedy
        # niepoliczalna i trzeba spasc do stopy referencyjnej NBP + 2%.
        margins_path = tmp_path / "edo.csv"
        margins_path.write_text(
            "issuance_month,first_year_rate_pct,margin_pct\n2026-01,,2.0\n", encoding="utf-8"
        )
        nbp_path = tmp_path / "nbp.csv"
        nbp_path.write_text("effective_from,reference_rate_pct\n2025-01-01,4.00\n", encoding="utf-8")

        fake_cpi = pd.DataFrame({"cpi_prev_year_100": [103.6]}, index=pd.Index([2025], name="year"))
        with patch("src.data_loader.fetch_gus_cpi", return_value=fake_cpi):
            result = build_edo_reference_rate_monthly(margins_path, nbp_path)

        expected_monthly = annualize_to_monthly(0.04 + 0.02)
        assert result.loc[pd.Period("2026-01", freq="M"), "edo_reference_monthly_return"] == pytest.approx(
            expected_monthly
        )
